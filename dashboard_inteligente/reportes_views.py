import json
import logging
from django.views import View
from django.http import HttpResponse, JsonResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import datetime
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from dashboard_inteligente.views import GenerarPrediccionesView
from productos.models import Producto
from ventas_carrito.models import Venta

logger = logging.getLogger(__name__)


@method_decorator(csrf_exempt, name='dispatch')
class ExportarDashboardVentasView(View):
    """Exportar reporte del Dashboard de Ventas en PDF o Excel"""
    
    def get(self, request):
        """Exportar dashboard de ventas"""
        try:
            if not request.session.get('is_authenticated'):
                return JsonResponse({
                    'success': False,
                    'message': 'Debe iniciar sesión'
                }, status=401)
            
            formato = request.GET.get('formato', 'pdf')  # pdf o excel
            periodo = request.GET.get('periodo', '12')  # meses a incluir
            
            # Obtener datos del dashboard directamente (SIEMPRE ACTUALIZADOS)
            from ventas_carrito.historial_views import DashboardStatsView
            stats_view = DashboardStatsView()
            stats_response = stats_view.get(request)
            
            if stats_response.status_code != 200:
                return JsonResponse({
                    'success': False,
                    'message': 'Error al obtener datos del dashboard'
                }, status=500)
            
            stats_data = json.loads(stats_response.content)
            
            # Obtener datos adicionales para el reporte detallado
            from ventas_carrito.models import Venta, DetalleVenta
            from productos.models import Producto, Categoria
            from autenticacion_usuarios.models import Cliente
            from django.db.models import Sum, Count, Avg, Q
            from datetime import timedelta
            
            ahora = timezone.now()
            inicio_mes_actual = ahora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            inicio_mes_anterior = (inicio_mes_actual - timedelta(days=1)).replace(day=1)
            
            # Ventas por categoría
            ventas_categoria = DetalleVenta.objects.filter(
                venta__fecha_venta__gte=inicio_mes_actual,
                venta__estado='completada',
                producto__categoria__isnull=False
            ).values('producto__categoria__nombre').annotate(
                total_ventas=Sum('subtotal'),
                cantidad_vendida=Sum('cantidad'),
                num_ventas=Count('venta', distinct=True)
            ).order_by('-total_ventas')[:10]
            
            # Clientes más activos
            clientes_activos = Venta.objects.filter(
                fecha_venta__gte=inicio_mes_actual,
                estado='completada'
            ).values(
                'cliente__id__nombre',
                'cliente__id__apellido',
                'cliente__id__email'
            ).annotate(
                total_compras=Sum('total'),
                num_compras=Count('id_venta')
            ).order_by('-total_compras')[:10]
            
            # Agregar datos adicionales al stats_data
            stats_data['ventas_por_categoria'] = [
                {
                    'categoria': item['producto__categoria__nombre'],
                    'total_ventas': float(item['total_ventas']),
                    'cantidad_vendida': item['cantidad_vendida'],
                    'num_ventas': item['num_ventas']
                }
                for item in ventas_categoria
            ]
            
            stats_data['clientes_activos'] = [
                {
                    'nombre': f"{item['cliente__id__nombre']} {item['cliente__id__apellido'] or ''}".strip(),
                    'email': item['cliente__id__email'],
                    'total_compras': float(item['total_compras']),
                    'num_compras': item['num_compras']
                }
                for item in clientes_activos
            ]
            
            if formato == 'pdf':
                return self._generar_pdf(stats_data, periodo)
            elif formato == 'excel':
                return self._generar_excel(stats_data, periodo)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Formato no soportado. Use pdf o excel'
                }, status=400)
                
        except Exception as e:
            logger.error(f"Error en ExportarDashboardVentasView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar reporte: {str(e)}'
            }, status=500)
    
    def _generar_pdf(self, stats_data, periodo):
        """Generar PDF del dashboard de ventas"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        styles = getSampleStyleSheet()
        
        # Estilos personalizados
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#0066FF'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        # Título
        story.append(Paragraph("Reporte de Dashboard de Ventas", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Información del reporte
        fecha_actual = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        info_data = [
            ['Fecha de Generación:', fecha_actual],
            ['Período Analizado:', f'Últimos {periodo} meses'],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1F2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Estadísticas principales
        if stats_data.get('success') and stats_data.get('stats'):
            stats = stats_data['stats']
            
            story.append(Paragraph("Estadísticas Principales", heading_style))
            
            stats_data_table = [
                ['Métrica', 'Valor', 'Cambio %', 'Tendencia'],
            ]
            
            if stats.get('ventas_mes'):
                stats_data_table.append([
                    'Ventas del Mes',
                    f"Bs. {stats['ventas_mes']['value']:,.2f}",
                    f"{stats['ventas_mes']['change']:+.1f}%",
                    stats['ventas_mes']['trend'].upper()
                ])
            
            if stats.get('total_pedidos'):
                stats_data_table.append([
                    'Total Pedidos',
                    str(stats['total_pedidos']['value']),
                    f"{stats['total_pedidos']['change']:+.1f}%",
                    stats['total_pedidos']['trend'].upper()
                ])
            
            if stats.get('nuevos_clientes'):
                stats_data_table.append([
                    'Nuevos Clientes',
                    str(stats['nuevos_clientes']['value']),
                    f"{stats['nuevos_clientes']['change']:+.1f}%",
                    stats['nuevos_clientes']['trend'].upper()
                ])
            
            if stats.get('productos_activos'):
                stats_data_table.append([
                    'Productos Activos',
                    str(stats['productos_activos']['value']),
                    f"{stats['productos_activos']['change']:+.1f}%",
                    stats['productos_activos']['trend'].upper()
                ])
            
            stats_table = Table(stats_data_table, colWidths=[2*inch, 2*inch, 1.5*inch, 1.5*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0066FF')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            story.append(stats_table)
            story.append(Spacer(1, 0.3*inch))
            
            # Ventas mensuales
            if stats.get('ventas_mensuales'):
                story.append(Paragraph("Ventas Mensuales", heading_style))
                
                ventas_mensuales = stats['ventas_mensuales']
                labels = ventas_mensuales.get('labels', [])
                values = ventas_mensuales.get('values', [])
                
                if labels and values:
                    ventas_table_data = [['Mes', 'Ventas (Bs.)']]
                    for i, label in enumerate(labels):
                        if i < len(values):
                            ventas_table_data.append([
                                label,
                                f"Bs. {values[i]:,.2f}"
                            ])
                    
                    ventas_table = Table(ventas_table_data, colWidths=[3*inch, 3*inch])
                    ventas_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                    ]))
                    story.append(ventas_table)
                    story.append(Spacer(1, 0.3*inch))
            
            # Productos top
            productos_top = stats_data.get('top_products') or stats.get('productos_top') or []
            if productos_top:
                story.append(Paragraph("Productos Más Vendidos (Top 15)", heading_style))
                
                # Normalizar formato de productos
                productos_normalizados = []
                for prod in productos_top[:15]:
                    if isinstance(prod, dict):
                        productos_normalizados.append({
                            'nombre': prod.get('name') or prod.get('nombre', 'N/A'),
                            'cantidad': prod.get('sales') or prod.get('cantidad', 0),
                            'total': prod.get('revenue') or prod.get('total', 0)
                        })
                
                productos_top = productos_normalizados
                productos_table_data = [['#', 'Producto', 'Unidades Vendidas', 'Total (Bs.)', 'Promedio/Unidad']]
                
                for idx, prod in enumerate(productos_top, 1):
                    cantidad = prod.get('cantidad', 0)
                    total = prod.get('total', 0)
                    promedio = total / cantidad if cantidad > 0 else 0
                    productos_table_data.append([
                        str(idx),
                        prod.get('nombre', 'N/A')[:40],  # Limitar longitud
                        str(cantidad),
                        f"Bs. {total:,.2f}",
                        f"Bs. {promedio:,.2f}"
                    ])
                
                productos_table = Table(productos_table_data, colWidths=[0.5*inch, 3*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                productos_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F59E0B')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ]))
                story.append(productos_table)
                story.append(Spacer(1, 0.3*inch))
            
            # Clientes más activos
            if stats_data.get('clientes_activos'):
                story.append(PageBreak())
                story.append(Paragraph("Clientes Más Activos (Top 10)", heading_style))
                
                clientes_table_data = [['#', 'Cliente', 'Email', 'Total Compras (Bs.)', 'N° Compras']]
                
                for idx, cliente in enumerate(stats_data['clientes_activos'], 1):
                    clientes_table_data.append([
                        str(idx),
                        cliente.get('nombre', 'N/A')[:30],
                        cliente.get('email', 'N/A')[:30],
                        f"Bs. {cliente.get('total_compras', 0):,.2f}",
                        str(cliente.get('num_compras', 0))
                    ])
                
                clientes_table = Table(clientes_table_data, colWidths=[0.5*inch, 2*inch, 2*inch, 1.5*inch, 1*inch])
                clientes_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Columna #
                    ('ALIGN', (1, 0), (2, -1), 'LEFT'),    # Columnas Cliente y Email
                    ('ALIGN', (3, 0), (-1, -1), 'CENTER'), # Resto de columnas
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ]))
                story.append(clientes_table)
                story.append(Spacer(1, 0.3*inch))
        
        # Pie de página
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(
            f"<i>Reporte generado el {fecha_actual} - SmartSales365</i>",
            styles['Normal']
        ))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="dashboard_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _generar_excel(self, stats_data, periodo):
        """Generar Excel del dashboard de ventas"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Ventas"
        
        # Título
        ws['A1'] = "Reporte de Dashboard de Ventas"
        ws['A1'].font = Font(bold=True, size=18, color="0066FF")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Información del reporte
        fecha_actual = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        ws['A3'] = 'Fecha de Generación:'
        ws['B3'] = fecha_actual
        ws['A4'] = 'Período Analizado:'
        ws['B4'] = f'Últimos {periodo} meses'
        
        row = 6
        
        if stats_data.get('success') and stats_data.get('stats'):
            stats = stats_data['stats']
            
            # Estadísticas principales
            ws[f'A{row}'] = 'Estadísticas Principales'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
            row += 1
            
            headers = ['Métrica', 'Valor', 'Cambio %', 'Tendencia']
            ws.append(headers)
            
            # Estilo para encabezados
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066FF", end_color="0066FF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Datos de estadísticas
            if stats.get('ventas_mes'):
                ws.append([
                    'Ventas del Mes',
                    f"Bs. {stats['ventas_mes']['value']:,.2f}",
                    f"{stats['ventas_mes']['change']:+.1f}%",
                    stats['ventas_mes']['trend'].upper()
                ])
                row += 1
            
            if stats.get('total_pedidos'):
                ws.append([
                    'Total Pedidos',
                    stats['total_pedidos']['value'],
                    f"{stats['total_pedidos']['change']:+.1f}%",
                    stats['total_pedidos']['trend'].upper()
                ])
                row += 1
            
            if stats.get('nuevos_clientes'):
                ws.append([
                    'Nuevos Clientes',
                    stats['nuevos_clientes']['value'],
                    f"{stats['nuevos_clientes']['change']:+.1f}%",
                    stats['nuevos_clientes']['trend'].upper()
                ])
                row += 1
            
            if stats.get('productos_activos'):
                ws.append([
                    'Productos Activos',
                    stats['productos_activos']['value'],
                    f"{stats['productos_activos']['change']:+.1f}%",
                    stats['productos_activos']['trend'].upper()
                ])
                row += 1
            
            row += 2
            
            # Ventas mensuales
            if stats.get('ventas_mensuales'):
                ws[f'A{row}'] = 'Ventas Mensuales'
                ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
                row += 1
                
                ventas_mensuales = stats['ventas_mensuales']
                labels = ventas_mensuales.get('labels', [])
                values = ventas_mensuales.get('values', [])
                
                ws.append(['Mes', 'Ventas (Bs.)'])
                # Estilo encabezado
                for col in range(1, 3):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for i, label in enumerate(labels):
                    if i < len(values):
                        ws.append([label, values[i]])
                        row += 1
                
                # Crear gráfico de barras
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                chart.title = "Ventas Mensuales"
                chart.y_axis.title = 'Ventas (Bs.)'
                chart.x_axis.title = 'Mes'
                
                data = Reference(ws, min_col=2, min_row=row-len(labels), max_row=row-1)
                cats = Reference(ws, min_col=1, min_row=row-len(labels), max_row=row-1)
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(cats)
                
                ws.add_chart(chart, f"E{row-len(labels)}")
                row += len(labels) + 5
            
            # Ventas por categoría
            if stats_data.get('ventas_por_categoria'):
                row += 2
                ws[f'A{row}'] = 'Ventas por Categoría'
                ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
                row += 1
                
                ws.append(['Categoría', 'Total Ventas (Bs.)', 'Cantidad Vendida', 'N° Ventas'])
                
                # Estilo encabezado
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for cat in stats_data['ventas_por_categoria']:
                    ws.append([
                        cat.get('categoria', 'N/A'),
                        cat.get('total_ventas', 0),
                        cat.get('cantidad_vendida', 0),
                        cat.get('num_ventas', 0)
                    ])
                    row += 1
                
                # Crear gráfico de barras para categorías
                if len(stats_data['ventas_por_categoria']) > 0:
                    chart_cat = BarChart()
                    chart_cat.type = "col"
                    chart_cat.style = 10
                    chart_cat.title = "Ventas por Categoría"
                    chart_cat.y_axis.title = 'Ventas (Bs.)'
                    chart_cat.x_axis.title = 'Categoría'
                    
                    data_cat = Reference(ws, min_col=2, min_row=row-len(stats_data['ventas_por_categoria']), max_row=row-1)
                    cats_cat = Reference(ws, min_col=1, min_row=row-len(stats_data['ventas_por_categoria']), max_row=row-1)
                    chart_cat.add_data(data_cat, titles_from_data=False)
                    chart_cat.set_categories(cats_cat)
                    
                    ws.add_chart(chart_cat, f"F{row-len(stats_data['ventas_por_categoria'])}")
                    row += len(stats_data['ventas_por_categoria']) + 5
            
            # Productos top
            productos_top = stats_data.get('top_products') or stats.get('productos_top') or []
            if productos_top:
                row += 2
                ws[f'A{row}'] = 'Productos Más Vendidos (Top 15)'
                ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
                row += 1
                
                # Normalizar formato de productos
                productos_normalizados = []
                for prod in productos_top[:15]:
                    if isinstance(prod, dict):
                        productos_normalizados.append({
                            'nombre': prod.get('name') or prod.get('nombre', 'N/A'),
                            'cantidad': prod.get('sales') or prod.get('cantidad', 0),
                            'total': prod.get('revenue') or prod.get('total', 0)
                        })
                
                productos_top = productos_normalizados
                ws.append(['#', 'Producto', 'Unidades Vendidas', 'Total (Bs.)', 'Promedio/Unidad'])
                
                # Estilo encabezado
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for idx, prod in enumerate(productos_top, 1):
                    cantidad = prod.get('cantidad', 0)
                    total = prod.get('total', 0)
                    promedio = total / cantidad if cantidad > 0 else 0
                    ws.append([
                        idx,
                        prod.get('nombre', 'N/A'),
                        cantidad,
                        total,
                        promedio
                    ])
                    row += 1
                
                # Crear gráfico de barras para productos
                if len(productos_top) > 0:
                    chart_prod = BarChart()
                    chart_prod.type = "col"
                    chart_prod.style = 10
                    chart_prod.title = "Top Productos por Ventas"
                    chart_prod.y_axis.title = 'Total (Bs.)'
                    chart_prod.x_axis.title = 'Producto'
                    
                    data_prod = Reference(ws, min_col=4, min_row=row-len(productos_top), max_row=row-1)
                    cats_prod = Reference(ws, min_col=2, min_row=row-len(productos_top), max_row=row-1)
                    chart_prod.add_data(data_prod, titles_from_data=False)
                    chart_prod.set_categories(cats_prod)
                    
                    ws.add_chart(chart_prod, f"G{row-len(productos_top)}")
                    row += len(productos_top) + 5
            
            # Clientes más activos
            if stats_data.get('clientes_activos'):
                row += 2
                ws[f'A{row}'] = 'Clientes Más Activos (Top 10)'
                ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
                row += 1
                
                ws.append(['#', 'Cliente', 'Email', 'Total Compras (Bs.)', 'N° Compras'])
                
                # Estilo encabezado
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for idx, cliente in enumerate(stats_data['clientes_activos'], 1):
                    ws.append([
                        idx,
                        cliente.get('nombre', 'N/A'),
                        cliente.get('email', 'N/A'),
                        cliente.get('total_compras', 0),
                        cliente.get('num_compras', 0)
                    ])
                    row += 1
                
                # Crear gráfico de barras para clientes
                if len(stats_data['clientes_activos']) > 0:
                    chart_cli = BarChart()
                    chart_cli.type = "col"
                    chart_cli.style = 10
                    chart_cli.title = "Top Clientes por Compras"
                    chart_cli.y_axis.title = 'Total Compras (Bs.)'
                    chart_cli.x_axis.title = 'Cliente'
                    
                    data_cli = Reference(ws, min_col=4, min_row=row-len(stats_data['clientes_activos']), max_row=row-1)
                    cats_cli = Reference(ws, min_col=2, min_row=row-len(stats_data['clientes_activos']), max_row=row-1)
                    chart_cli.add_data(data_cli, titles_from_data=False)
                    chart_cli.set_categories(cats_cli)
                    
                    ws.add_chart(chart_cli, f"G{row-len(stats_data['clientes_activos'])}")
                    row += len(stats_data['clientes_activos']) + 5
            
            # Análisis de tendencias
            if stats.get('ventas_mensuales'):
                row += 2
                ws[f'A{row}'] = 'Análisis de Tendencias Mensuales'
                ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
                row += 1
                
                ventas_mensuales = stats['ventas_mensuales']
                labels = ventas_mensuales.get('labels', [])
                values = ventas_mensuales.get('values', [])
                
                ws.append(['Mes', 'Ventas (Bs.)', 'Crecimiento %', 'Tendencia'])
                
                # Estilo encabezado
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for i, label in enumerate(labels):
                    if i < len(values):
                        valor_actual = values[i]
                        crecimiento = 0
                        if i > 0:
                            valor_anterior = values[i-1]
                            if valor_anterior > 0:
                                crecimiento = ((valor_actual - valor_anterior) / valor_anterior) * 100
                        
                        tendencia = '↑' if crecimiento > 0 else '↓' if crecimiento < 0 else '→'
                        ws.append([label, valor_actual, f"{crecimiento:+.1f}%", tendencia])
                        row += 1
                
                # Crear gráfico de línea para tendencias
                if len(labels) > 0:
                    chart_tend = LineChart()
                    chart_tend.title = "Tendencia de Ventas Mensuales"
                    chart_tend.y_axis.title = 'Ventas (Bs.)'
                    chart_tend.x_axis.title = 'Mes'
                    
                    data_tend = Reference(ws, min_col=2, min_row=row-len(labels), max_row=row-1)
                    cats_tend = Reference(ws, min_col=1, min_row=row-len(labels), max_row=row-1)
                    chart_tend.add_data(data_tend, titles_from_data=False)
                    chart_tend.set_categories(cats_tend)
                    
                    ws.add_chart(chart_tend, f"F{row-len(labels)}")
                    row += len(labels) + 5
        
        # Ajustar ancho de columnas dinámicamente
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 25
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="dashboard_ventas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response


@method_decorator(csrf_exempt, name='dispatch')
class ExportarPrediccionesView(View):
    """Exportar reporte de Predicciones de IA en PDF o Excel"""
    
    def get(self, request):
        """Exportar predicciones"""
        try:
            if not request.session.get('is_authenticated'):
                return JsonResponse({
                    'success': False,
                    'message': 'Debe iniciar sesión'
                }, status=401)
            
            formato = request.GET.get('formato', 'pdf')  # pdf o excel
            ids_predicciones = request.GET.get('ids', None)  # IDs específicos de predicciones a exportar
            
            # Obtener predicciones - usar listar predicciones
            from reportes_dinamicos.models import PrediccionVenta
            from productos.models import Categoria
            
            # Si se especifican IDs, obtener solo esas predicciones
            if ids_predicciones:
                try:
                    ids_list = [int(id_str.strip()) for id_str in ids_predicciones.split(',') if id_str.strip()]
                    predicciones_queryset = PrediccionVenta.objects.filter(id_prediccion__in=ids_list).order_by('-fecha_prediccion')
                except (ValueError, TypeError):
                    # Si hay error parseando IDs, obtener todas
                    predicciones_queryset = PrediccionVenta.objects.all().order_by('-fecha_prediccion')[:100]
            else:
                # Obtener todas las predicciones recientes
                predicciones_queryset = PrediccionVenta.objects.all().order_by('-fecha_prediccion')[:100]
            
            predicciones_list = []
            for pred in predicciones_queryset:
                categoria_data = None
                if pred.categoria:
                    categoria_data = {
                        'id_categoria': pred.categoria.id_categoria,
                        'nombre': pred.categoria.nombre
                    }
                
                predicciones_list.append({
                    'id': pred.id_prediccion,
                    'fecha_prediccion': pred.fecha_prediccion.strftime('%Y-%m-%d') if pred.fecha_prediccion else None,
                    'valor_predicho': float(pred.valor_predicho) if pred.valor_predicho else 0,
                    'confianza': float(pred.confianza) if pred.confianza else 0,
                    'categoria': categoria_data
                })
            
            # Calcular resumen
            total_predicciones = len(predicciones_list)
            total_valor_predicho = sum(p['valor_predicho'] for p in predicciones_list)
            confianza_promedio = sum(p['confianza'] for p in predicciones_list) / total_predicciones if total_predicciones > 0 else 0
            
            # Calcular tendencias basadas en datos históricos reales
            from ventas_carrito.models import Venta
            from django.db.models import Sum
            from datetime import timedelta
            
            fecha_actual = timezone.now()
            fecha_inicio_historico = fecha_actual - timedelta(days=90)  # Últimos 3 meses
            
            ventas_recientes = Venta.objects.filter(
                estado='completada',
                fecha_venta__gte=fecha_inicio_historico
            )
            
            total_ventas_recientes_raw = ventas_recientes.aggregate(total=Sum('total'))['total'] or 0
            total_ventas_recientes = float(total_ventas_recientes_raw) if total_ventas_recientes_raw else 0.0
            promedio_mensual_historico = float(total_ventas_recientes / 3) if total_ventas_recientes > 0 else 0.0
            
            # Calcular factor de crecimiento (últimos 30 días vs 30 días anteriores)
            ultimos_30_dias = fecha_actual - timedelta(days=30)
            anteriores_30_dias = ultimos_30_dias - timedelta(days=30)
            
            ventas_ultimos_30_raw = ventas_recientes.filter(
                fecha_venta__gte=ultimos_30_dias
            ).aggregate(total=Sum('total'))['total'] or 0
            ventas_ultimos_30 = float(ventas_ultimos_30_raw) if ventas_ultimos_30_raw else 0.0
            
            ventas_anteriores_30_raw = ventas_recientes.filter(
                fecha_venta__gte=anteriores_30_dias,
                fecha_venta__lt=ultimos_30_dias
            ).aggregate(total=Sum('total'))['total'] or 0
            ventas_anteriores_30 = float(ventas_anteriores_30_raw) if ventas_anteriores_30_raw else 0.0
            
            if ventas_anteriores_30 > 0:
                factor_crecimiento = float((ventas_ultimos_30 - ventas_anteriores_30) / ventas_anteriores_30) * 100
            else:
                factor_crecimiento = 0.0
            
            predicciones_data = {
                'success': True,
                'predicciones': predicciones_list,
                'resumen': {
                    'total_predicciones': total_predicciones,
                    'total_valor_predicho': total_valor_predicho,
                    'confianza_promedio': confianza_promedio,
                    'tendencias': {
                        'factor_crecimiento': round(factor_crecimiento, 2),
                        'promedio_mensual_historico': round(promedio_mensual_historico, 2),
                        'ventas_ultimos_30_dias': round(ventas_ultimos_30, 2),
                        'ventas_anteriores_30_dias': round(ventas_anteriores_30, 2)
                    }
                }
            }
            
            # Obtener estado del modelo
            from dashboard_inteligente.views import EstadoModeloView
            modelo_view = EstadoModeloView()
            modelo_response = modelo_view.get(request)
            modelo_data = json.loads(modelo_response.content) if modelo_response.status_code == 200 else None
            
            if formato == 'pdf':
                return self._generar_pdf(predicciones_data, modelo_data)
            elif formato == 'excel':
                return self._generar_excel(predicciones_data, modelo_data)
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Formato no soportado. Use pdf o excel'
                }, status=400)
                
        except Exception as e:
            logger.error(f"Error en ExportarPrediccionesView: {str(e)}", exc_info=True)
            return JsonResponse({
                'success': False,
                'message': f'Error al exportar reporte: {str(e)}'
            }, status=500)
    
    def _generar_pdf(self, predicciones_data, modelo_data):
        """Generar PDF de predicciones con detalles mejorados"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#8B5CF6'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#1F2937'),
            spaceAfter=12,
            spaceBefore=20,
            fontName='Helvetica-Bold'
        )
        
        subheading_style = ParagraphStyle(
            'CustomSubHeading',
            parent=styles['Heading3'],
            fontSize=14,
            textColor=colors.HexColor('#4B5563'),
            spaceAfter=8,
            spaceBefore=12,
            fontName='Helvetica-Bold'
        )
        
        # Título
        story.append(Paragraph("Reporte de Predicciones de IA", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        fecha_actual = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        info_data = [
            ['Fecha de Generación:', fecha_actual],
        ]
        
        if modelo_data and modelo_data.get('modelo'):
            modelo = modelo_data['modelo']
            info_data.append(['Modelo de IA:', f"{modelo.get('nombre', 'N/A')} v{modelo.get('version', 'N/A')}"])
            info_data.append(['Estado del Modelo:', modelo.get('estado', 'N/A').upper()])
            if modelo.get('r2_score'):
                info_data.append(['R² Score (Calidad):', f"{modelo['r2_score']:.3f}"])
            if modelo.get('registros_entrenamiento'):
                info_data.append(['Registros de Entrenamiento:', f"{modelo['registros_entrenamiento']:,}"])
        
        info_table = Table(info_data, colWidths=[2.5*inch, 3.5*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F3F4F6')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1F2937')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Resumen de predicciones
        if predicciones_data.get('success') and predicciones_data.get('resumen'):
            resumen = predicciones_data['resumen']
            story.append(Paragraph("Resumen Ejecutivo de Predicciones", heading_style))
            
            resumen_data = [
                ['Métrica', 'Valor'],
                ['Total de Predicciones:', str(resumen.get('total_predicciones', 0))],
                ['Total Valor Predicho:', f"Bs. {resumen.get('total_valor_predicho', 0):,.2f}"],
                ['Confianza Promedio:', f"{resumen.get('confianza_promedio', 0) * 100:.1f}%"],
            ]
            
            if resumen.get('tendencias'):
                tendencias = resumen['tendencias']
                resumen_data.append(['Factor de Crecimiento:', f"{tendencias.get('factor_crecimiento', 0):+.1f}%"])
                resumen_data.append(['Promedio Mensual Histórico:', f"Bs. {tendencias.get('promedio_mensual_historico', 0):,.2f}"])
                if tendencias.get('ventas_ultimos_30_dias'):
                    resumen_data.append(['Ventas Últimos 30 Días:', f"Bs. {tendencias.get('ventas_ultimos_30_dias', 0):,.2f}"])
                if tendencias.get('ventas_anteriores_30_dias'):
                    resumen_data.append(['Ventas 30 Días Anteriores:', f"Bs. {tendencias.get('ventas_anteriores_30_dias', 0):,.2f}"])
            
            resumen_table = Table(resumen_data, colWidths=[3*inch, 3*inch])
            resumen_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 11),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F3F4F6')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            story.append(resumen_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Análisis por categoría (si hay múltiples categorías)
        if predicciones_data.get('success') and predicciones_data.get('predicciones'):
            predicciones = predicciones_data['predicciones']
            categorias_dict = {}
            for pred in predicciones:
                cat_nombre = pred.get('categoria', {}).get('nombre', 'General') if pred.get('categoria') else 'General'
                if cat_nombre not in categorias_dict:
                    categorias_dict[cat_nombre] = {'total': 0, 'count': 0, 'confianza_promedio': 0}
                categorias_dict[cat_nombre]['total'] += pred.get('valor_predicho', 0)
                categorias_dict[cat_nombre]['count'] += 1
                categorias_dict[cat_nombre]['confianza_promedio'] += pred.get('confianza', 0)
            
            if len(categorias_dict) > 1:
                story.append(PageBreak())
                story.append(Paragraph("Análisis por Categoría", heading_style))
                
                categoria_table_data = [['Categoría', 'Total Predicho (Bs.)', 'N° Predicciones', 'Confianza Promedio']]
                for cat_nombre, datos in sorted(categorias_dict.items(), key=lambda x: x[1]['total'], reverse=True):
                    confianza_avg = (datos['confianza_promedio'] / datos['count']) * 100 if datos['count'] > 0 else 0
                    categoria_table_data.append([
                        cat_nombre,
                        f"{datos['total']:,.2f}",
                        str(datos['count']),
                        f"{confianza_avg:.1f}%"
                    ])
                
                categoria_table = Table(categoria_table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                categoria_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                    ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ]))
                story.append(categoria_table)
                story.append(Spacer(1, 0.3*inch))
        
        # Lista detallada de predicciones
        if predicciones_data.get('success') and predicciones_data.get('predicciones'):
            predicciones = predicciones_data['predicciones']
            # Verificar si categorias_dict existe (se crea en el bloque anterior)
            categorias_dict = {}
            if predicciones:
                for pred in predicciones:
                    cat_nombre = pred.get('categoria', {}).get('nombre', 'General') if pred.get('categoria') else 'General'
                    if cat_nombre not in categorias_dict:
                        categorias_dict[cat_nombre] = {'total': 0, 'count': 0, 'confianza_promedio': 0}
            
            story.append(PageBreak() if len(categorias_dict) > 1 else Spacer(1, 0.3*inch))
            story.append(Paragraph("Predicciones Detalladas", heading_style))
            
            predicciones_table_data = [['Fecha', 'Valor Predicho (Bs.)', 'Confianza', 'Categoría']]
            
            # Ordenar por fecha
            predicciones_ordenadas = sorted(predicciones, key=lambda x: x.get('fecha_prediccion', ''))
            
            for pred in predicciones_ordenadas[:100]:  # Limitar a 100 para PDF
                predicciones_table_data.append([
                    pred.get('fecha_prediccion', 'N/A')[:10] if pred.get('fecha_prediccion') else 'N/A',
                    f"{pred.get('valor_predicho', 0):,.2f}",
                    f"{pred.get('confianza', 0) * 100:.1f}%",
                    pred.get('categoria', {}).get('nombre', 'General') if pred.get('categoria') else 'General'
                ])
            
            predicciones_table = Table(predicciones_table_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
            predicciones_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ]))
            story.append(predicciones_table)
            
            if len(predicciones) > 100:
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph(
                    f"<i>Nota: Se muestran 100 de {len(predicciones)} predicciones totales.</i>",
                    styles['Normal']
                ))
        
        story.append(Spacer(1, 0.5*inch))
        story.append(Paragraph(
            f"<i>Reporte generado el {fecha_actual} - SmartSales365 - Sistema de Predicciones con IA</i>",
            styles['Normal']
        ))
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="predicciones_ia_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        return response
    
    def _generar_excel(self, predicciones_data, modelo_data):
        """Generar Excel de predicciones con detalles mejorados y gráficos"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Predicciones IA"
        
        # Título
        ws['A1'] = "Reporte de Predicciones de IA"
        ws['A1'].font = Font(bold=True, size=18, color="8B5CF6")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        fecha_actual = timezone.now().strftime('%d/%m/%Y %H:%M:%S')
        ws['A3'] = 'Fecha de Generación:'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = fecha_actual
        
        row = 5
        
        # Información del modelo
        if modelo_data and modelo_data.get('modelo'):
            modelo = modelo_data['modelo']
            ws[f'A{row}'] = 'Información del Modelo'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
            row += 1
            
            ws.append(['Campo', 'Valor'])
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            ws.append(['Nombre', modelo.get('nombre', 'N/A')])
            row += 1
            ws.append(['Versión', modelo.get('version', 'N/A')])
            row += 1
            ws.append(['Estado', modelo.get('estado', 'N/A').upper()])
            row += 1
            if modelo.get('r2_score'):
                ws.append(['R² Score (Calidad)', f"{modelo['r2_score']:.3f}"])
                row += 1
            if modelo.get('registros_entrenamiento'):
                ws.append(['Registros de Entrenamiento', f"{modelo['registros_entrenamiento']:,}"])
                row += 1
            row += 2
        
        # Resumen ejecutivo
        if predicciones_data.get('success') and predicciones_data.get('resumen'):
            resumen = predicciones_data['resumen']
            ws[f'A{row}'] = 'Resumen Ejecutivo de Predicciones'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
            row += 1
            
            ws.append(['Métrica', 'Valor'])
            for col in range(1, 3):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            ws.append(['Total Predicciones', resumen.get('total_predicciones', 0)])
            row += 1
            ws.append(['Total Valor Predicho', f"Bs. {resumen.get('total_valor_predicho', 0):,.2f}"])
            row += 1
            ws.append(['Confianza Promedio', f"{resumen.get('confianza_promedio', 0) * 100:.1f}%"])
            row += 1
            
            if resumen.get('tendencias'):
                tendencias = resumen['tendencias']
                ws.append(['Factor de Crecimiento', f"{tendencias.get('factor_crecimiento', 0):+.1f}%"])
                row += 1
                ws.append(['Promedio Mensual Histórico', f"Bs. {tendencias.get('promedio_mensual_historico', 0):,.2f}"])
                row += 1
                if tendencias.get('ventas_ultimos_30_dias'):
                    ws.append(['Ventas Últimos 30 Días', f"Bs. {tendencias.get('ventas_ultimos_30_dias', 0):,.2f}"])
                    row += 1
                if tendencias.get('ventas_anteriores_30_dias'):
                    ws.append(['Ventas 30 Días Anteriores', f"Bs. {tendencias.get('ventas_anteriores_30_dias', 0):,.2f}"])
                    row += 1
            
            row += 2
        
        # Análisis por categoría
        if predicciones_data.get('success') and predicciones_data.get('predicciones'):
            predicciones = predicciones_data['predicciones']
            categorias_dict = {}
            for pred in predicciones:
                cat_nombre = pred.get('categoria', {}).get('nombre', 'General') if pred.get('categoria') else 'General'
                if cat_nombre not in categorias_dict:
                    categorias_dict[cat_nombre] = {'total': 0, 'count': 0, 'confianza_promedio': 0}
                categorias_dict[cat_nombre]['total'] += pred.get('valor_predicho', 0)
                categorias_dict[cat_nombre]['count'] += 1
                categorias_dict[cat_nombre]['confianza_promedio'] += pred.get('confianza', 0)
            
            if len(categorias_dict) > 1:
                ws[f'A{row}'] = 'Análisis por Categoría'
                ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
                row += 1
                
                ws.append(['Categoría', 'Total Predicho (Bs.)', 'N° Predicciones', 'Confianza Promedio (%)'])
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for cat_nombre, datos in sorted(categorias_dict.items(), key=lambda x: x[1]['total'], reverse=True):
                    confianza_avg = (datos['confianza_promedio'] / datos['count']) * 100 if datos['count'] > 0 else 0
                    ws.append([
                        cat_nombre,
                        datos['total'],
                        datos['count'],
                        round(confianza_avg, 1)
                    ])
                    row += 1
                
                # Gráfico de barras por categoría
                if len(categorias_dict) > 0:
                    chart_cat = BarChart()
                    chart_cat.title = "Predicciones por Categoría"
                    chart_cat.y_axis.title = 'Valor Predicho (Bs.)'
                    chart_cat.x_axis.title = 'Categoría'
                    
                    data_cat = Reference(ws, min_col=2, min_row=row-len(categorias_dict), max_row=row-1)
                    cats_ref = Reference(ws, min_col=1, min_row=row-len(categorias_dict), max_row=row-1)
                    chart_cat.add_data(data_cat, titles_from_data=False)
                    chart_cat.set_categories(cats_ref)
                    chart_cat.width = 12
                    chart_cat.height = 7
                    
                    ws.add_chart(chart_cat, f"F{row-len(categorias_dict)}")
                    row += 10
        
        # Predicciones detalladas
        if predicciones_data.get('success') and predicciones_data.get('predicciones'):
            predicciones = predicciones_data['predicciones']
            row += 2
            ws[f'A{row}'] = 'Predicciones Detalladas'
            ws[f'A{row}'].font = Font(bold=True, size=14, color="1F2937")
            row += 1
            
            ws.append(['Fecha', 'Valor Predicho (Bs.)', 'Confianza (%)', 'Categoría'])
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Ordenar por fecha
            predicciones_ordenadas = sorted(predicciones, key=lambda x: x.get('fecha_prediccion', ''))
            start_row = row
            
            for pred in predicciones_ordenadas:
                ws.append([
                    pred.get('fecha_prediccion', 'N/A')[:10] if pred.get('fecha_prediccion') else 'N/A',
                    pred.get('valor_predicho', 0),
                    round(pred.get('confianza', 0) * 100, 1),
                    pred.get('categoria', {}).get('nombre', 'General') if pred.get('categoria') else 'General'
                ])
                row += 1
            
            # Gráfico de línea - Evolución de predicciones
            if len(predicciones_ordenadas) > 0:
                chart = LineChart()
                chart.title = "Evolución de Predicciones en el Tiempo"
                chart.y_axis.title = 'Valor Predicho (Bs.)'
                chart.x_axis.title = 'Fecha'
                chart.width = 15
                chart.height = 8
                
                data = Reference(ws, min_col=2, min_row=start_row, max_row=row-1)
                cats = Reference(ws, min_col=1, min_row=start_row, max_row=row-1)
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(cats)
                
                ws.add_chart(chart, f"F{start_row}")
                
                # Gráfico de barras - Confianza promedio
                row_chart = row + 2
                chart_conf = BarChart()
                chart_conf.title = "Nivel de Confianza de Predicciones"
                chart_conf.y_axis.title = 'Confianza (%)'
                chart_conf.x_axis.title = 'Fecha'
                chart_conf.width = 15
                chart_conf.height = 8
                
                data_conf = Reference(ws, min_col=3, min_row=start_row, max_row=row-1)
                chart_conf.add_data(data_conf, titles_from_data=False)
                chart_conf.set_categories(cats)
                
                ws.add_chart(chart_conf, f"F{row_chart}")
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="predicciones_ia_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        return response

